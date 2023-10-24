import os
import csv
import pandas as pd
from openpyxl.styles import PatternFill


def has_only_pdf_urls(csv_file):
    with open(csv_file, 'r', newline='', encoding='utf-8') as file:
        reader = csv.reader(file)
        for row_number, row in enumerate(reader, start=1):
            # Skip the row if it's just a separator, remove this if necessary
            if row and '||' in row[0]:
                continue
            pdf_url = row[0]
            if not pdf_url.endswith('.pdf'):
                return False
        return True


def find_all_pdf_urls(csv_file):
    pdf_urls = []
    with open(csv_file, 'r', newline='', encoding='utf-8') as file:
        reader = csv.reader(file)
        for row_number, row in enumerate(reader, start=1):
            # Skip the row if it's just a separator, remove this if necessary
            if row and '||' in row[0]:
                continue
            pdf_url = row[0]
            if pdf_url.endswith('.pdf'):
                pdf_urls.append(pdf_url)
    return pdf_urls


if __name__ == '__main__':
    file_path = 'resources/QualityTest2.xlsx'
    file_dir = 'temp_files'
    excel_output = 'result.xlsx'
    df = pd.read_excel(file_path, sheet_name=0)
    num_rows = df.shape[0]
    num_cols = df.shape[1]

    with pd.ExcelWriter(excel_output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='result', index=False)
        workbook = writer.book
        sheet = writer.sheets['result']

        green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        blue_fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')

        index = 1
        for row in sheet.iter_rows(min_row=2, max_row=num_rows+1, max_col=num_cols):
            csv_file_path = os.path.join(file_dir, f"{index}.csv")
            if os.path.exists(csv_file_path):
                if os.path.getsize(csv_file_path) == 0:
                    cell.fill = red_fill
                elif has_only_pdf_urls(csv_file_path):
                    found_pdfs = find_all_pdf_urls(csv_file_path)
                    for cell in row:
                        cell.fill = yellow_fill
            else:
                for cell in row:
                    cell.fill = red_fill
            index += 1

        '''for i in range(1, num_rows + 1):
            csv_file_path = os.path.join(file_dir, f"{i}.csv")
            if os.path.exists(csv_file_path):
                if os.path.getsize(csv_file_path) == 0:
                    for cell in sheet[i + 1]:
                        cell.fill = red_fill
                elif has_only_pdf_urls(csv_file_path):
                    found_pdfs = find_all_pdf_urls(csv_file_path)
                    for cell in sheet[i + 1]:
                        cell.fill = yellow_fill
                    for index, value in enumerate(found_pdfs, num_cols+2):
                        sheet.cell(row=i + 1, column=index, value=value).fill = yellow_fill
                else:
                    for cell in sheet[i + 1]:
                        # Add the found data to cells in excel
                        cell.fill = green_fill
            else:
                print(f"Error: {csv_file_path} does not exist")
                for cell in sheet[i + 1]:
                    cell.fill = red_fill
                # Handle the error as required'''

        workbook.save(excel_output)
