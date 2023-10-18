from ExcelManip import excelmanip as em
from WebCrawler.crawlerDemo4 import WebCrawler
from urllib.parse import quote
import streamlit as st #for UI
import os #for UI
import shutil #for UI
import pandas as pd#for UI
import csv
from openpyxl.styles import PatternFill

def updateNumbers():
    nummer_found.markdown(f'RSK or E-Nummer found: {countNummer} ({countNummer / i * 100:.2f}%)')
    search_performed.markdown(f'Google Searches performed: {countGoogleSearch} ({countGoogleSearch / i * 100:.2f}%)')

def process_with_id(id_val, web_crawler, index):
    websites = [
        f'https://www.e-nummersok.se/sok?Query={id_val}',
        f'https://www.rskdatabasen.se/sok?Query={id_val}'
    ]
    for url in websites:
        web_crawler.crawl_website_with_depth(str(index), 0, start_url=url)

def has_only_pdf_urls(csv_file):
    with open(csv_file, 'r', newline='', encoding='utf-8') as file:
        reader = csv.reader(file)
        for row_number, row in enumerate(reader, start=1):
            # Skip the row if it's just a separator, remove this if necessary
            if row and '||' in row[0]:
                continue
            pdf_url = row[0]
            if not pdf_url.endswith('.pdf'):
                return False
        return True

def preProcess(file_path):
    # file_path = 'resources/100_items.xlsx'
    file_dir = 'temp_files'
    excel_output = 'result.xlsx'
    df = pd.read_excel(file_path, sheet_name=0)
    num_rows = len(df)

    with pd.ExcelWriter(excel_output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='result', index=False)
        workbook = writer.book
        sheet = writer.sheets['result']

        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

        for i in range(1, num_rows + 1):
            csv_file_path = os.path.join(file_dir, f"{i}.csv")
            if os.path.exists(csv_file_path):
                if os.path.getsize(csv_file_path) == 0:
                    for cell in sheet[i + 1]:
                        cell.fill = red_fill
                elif has_only_pdf_urls(csv_file_path):
                    for cell in sheet[i + 1]:
                        cell.fill = yellow_fill
            else:
                print(f"Error: {csv_file_path} does not exist")
                # Handle the error as required

        workbook.save(excel_output)

def process_without_id(data_dict, web_crawler, index):
    filtered_values = [value for value in data_dict.values() if value is not None]
    if filtered_values:
        search_query = " ".join(filtered_values)
        search_query = quote(search_query)
        google_search_url = f'https://www.google.com/search?q={search_query}'
        web_crawler.crawl_website_with_depth(str(index), 1, start_url=google_search_url)
    else:
        print("Dictionary has no valid values to perform a search.")

if 'files_processed' not in st.session_state:
    st.session_state.files_processed = False

# Note: using unsafe_allow_html can cause security problems if application is deployed on the web.
# belows Code section is used for the streamlit UI
st.markdown("""
    <style>
     @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@600&display=swap');
        .center {
            display: block;
            margin-left: auto;
            margin-right: auto;
            text-align: center;
            font-family: Poppins, sans-serif;
            font-size: 75px;
        }
        .subtitle{
            display: block;
            margin-left: auto;
            margin-right: auto;
            margin-bottom: 20px;
            text-align: center;
            font-family: Poppins, sans-serif;
            font-size: 40px;
        }
        .text{
            display: block;
            margin-left: auto;
            margin-right: auto;
            text-align: center;
            font-family: 'Poppins', sans-serif;
            font-size: 13px;
            font-weight: 100;
        }
    </style>
    """, unsafe_allow_html=True)
st.markdown("<h1 class='center'>SSG</h1>", unsafe_allow_html=True)
st.markdown("<h1 class='subtitle' 'center'>Data Enrichment Tool</h1>", unsafe_allow_html=True)

upload_excel = st.file_uploader("Choose an Excel file", type="xlsx")
if upload_excel:
    original_file_name = upload_excel.name.rsplit(".", 1)[0]
    df = pd.read_excel(upload_excel)
    num_rows = len(df)
    countNummer = 0
    countGoogleSearch = 0
    st.markdown(f'<p class="text">Number of Lines (Items): {num_rows}</p>', unsafe_allow_html=True)
    processed_text = st.markdown(f'<p class="text">Items processed: not yet started</p>', unsafe_allow_html=True)
    my_expander = st.expander('Click to expand for more information', expanded=False)
    nummer_found = my_expander.markdown('RSK or E-Nummer found: ...')
    search_performed = my_expander.markdown('Google Searches performed: ...')

    print(f"Number of Lines (Items): {num_rows}")
    del df
    if st.button("Start processing File"):
        progress_bar = st.progress(0)
        status_text = st.empty()
        processed_text.markdown(f'<p class="text">Items processed: 0</p>', unsafe_allow_html=True)
# following code section is mostly for Web Crawling
        excel = em.ExcelManip(upload_excel)
        data = excel.pre_process()
        wc = WebCrawler()
        i = 1

        for idx, dictionary in enumerate(data, start=1):
            if 'id' in dictionary and dictionary['id'] is not None:
                id_nr = dictionary['id']
                process_with_id(id_nr, wc, idx)
                process_without_id(dictionary, wc, idx)
                countNummer += 1
                updateNumbers()
            else:
                process_without_id(dictionary, wc, idx)
                countGoogleSearch += 1
                updateNumbers()

# next 4 code lines is for UI (update the progress bar, the number of processed Items, and the percentage number for the progress bar
            progress_percentage = (i/num_rows)*100
            progress_bar.progress(int(progress_percentage)) #update the progress bar once it is done with scraping
            processed_text.markdown(f'<p class="text">Items processed: {i}</p>', unsafe_allow_html=True) #update the number of items procesed
            status_text.markdown(f'<p class="text">Progress: {progress_percentage:.2f}%</p>', unsafe_allow_html=True) #update the percentage text
            i += 1
        preProcess(upload_excel)
        st.session_state.files_processed = True
        print("------Done------")
        wc.close()

        # here the Preprocess will happen

#code section below is for UI (Processing complete, prepare zip for download
        status_text.markdown("<p class='text'>Processing Complete!</p>", unsafe_allow_html=True)


    if st.session_state.files_processed:
        zip_name = "download.zip"
        if os.path.exists('temp_files'):
            shutil.make_archive(zip_name.replace('.zip', ''), 'zip', 'temp_files', )

        with open(zip_name, "rb") as file:
            st.download_button( label="Download Zipped CSV Files", data=file, file_name=f"Processed_{original_file_name}.zip", mime='application/zip', )
            # remove Files after Download
            # os.remove(zip_name)

        with open("result.xlsx", "rb") as file:
            btn = st.download_button(
                label="Download Result Excel",
                data=file,
                file_name=f"Processed_{original_file_name}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        if st.button("Clean and Restart"):
            # remove zipped File after Download
            os.remove(zip_name)

            # remove CSV Files in temp_files
            folder_path = 'temp_files'
            file_paths = [os.path.join(folder_path, file_name) for file_name in os.listdir(folder_path)]
            for file_path in file_paths:
                try:
                    os.remove(file_path)
                except Exception as e:
                    print(f"Error occurred while deleting file {file_path}: {str(e)}")

            # Delete all downloaded PDF Files in the directory
            folder_path = os.getcwd()
            file_paths = [os.path.join(folder_path, file_name) for file_name in os.listdir(folder_path)]
            for file_path in file_paths:
                try:
                    # Check if the file is a PDF by looking at the extension
                    if file_path.lower().endswith(".pdf"):
                        os.remove(file_path)
                        print(f"Deleted file: {file_path}")
                    else:
                        print(f"Skipped file: {file_path}")
                except Exception as e:
                    print(f"Error occurred while deleting file {file_path}: {str(e)}")

            #remove result.xlsx
            os.remove("result.xlsx")


            st.session_state.files_processed = False
            st.rerun()


