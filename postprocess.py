import os
import csv
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

GREEN_FILL = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
RED_FILL = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
BLUE_FILL = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')


class CSVHandler:
    """
    A class for handling CSV files and extracting information from them.

    Methods:
        has_only_pdf_urls(csv_file):
            Check if a CSV file contains only PDF URLs and return a boolean indicating this and a list of PDF URLs.

        find_all_pdf_urls(csv_file):
            Find and return all PDF URLs in a CSV file.
    """
    @staticmethod
    def has_only_pdf_urls(csv_file):
        """
        Check if a CSV file contains only PDF URLs and return a boolean indicating this and a list of PDF URLs.

        Args:
            csv_file (str): The path to the CSV file to be checked.
        Returns:
            tuple: A tuple containing a boolean (True if only PDF URLs, False otherwise) and a list of PDF URLs.
        """
        pdf_urls = CSVHandler.find_all_pdf_urls(csv_file)
        return all(url.endswith('.pdf') for url in pdf_urls), pdf_urls

    @staticmethod
    def find_all_pdf_urls(csv_file):
        """
        Find and return all PDF URLs in a CSV file.

        Args:
            csv_file (str): The path to the CSV file to be processed.

        Returns:
            list: A list of PDF URLs found in the CSV file.
        """
        try:
            with open(csv_file, 'r', newline='', encoding='utf-8') as file:
                return [row[0] for row in csv.reader(file) if row and not '||' in row[0]]
        except (FileNotFoundError, UnicodeDecodeError, csv.Error) as e:
            # Handle the exception(s) here
            print(f"An error occurred while processing the CSV file: {e}")
            return []


class ExcelProcessor:
    """
    A class for processing Excel files, generating reports, and integrating data from CSV files.

    Attributes:
        input_file (str): The path to the input Excel file.
        output_file (str): The path to the output Excel file.
        csv_directory (str): The directory where CSV files are stored.

    Methods:
        _fill_row(row, fill_color):
            Fill cells in a row with a specified fill color.

        _add_to_row(sheet, row_index, num_cols, found_pdfs):
            Add data to a row in an Excel sheet based on the presence of PDF URLs.

        _process_csv_files(sheet, num_rows, num_cols):
            Process CSV files associated with the Excel sheet and update cell fill colors accordingly.

        process_excel():
            Process the input Excel file, integrate data from CSV files, and generate an output Excel file.
    """

    def __init__(self, input_file, output_file, csv_directory):
        """
        Initialize the ExcelProcessor object.
        :param input_file: The path to the input Excel file.
        :param output_file: The path to the output Excel file.
        :param csv_directory: The directory where the CSV files are stored.
        """
        self.input_file = input_file
        self.output_file = output_file
        self.csv_directory = csv_directory

    def _fill_row(self, row, fill_color):
        for cell in row:
            cell.fill = fill_color

    def _add_to_row(self, sheet, row_index, num_cols, found_pdfs):
        for i, data_value in enumerate(found_pdfs, start=2):
            cell = sheet.cell(row=row_index + 1, column=num_cols + i)
            cell.value = data_value

    def _process_csv_files(self, sheet, num_rows, num_cols):
        for index, row in enumerate(sheet.iter_rows(min_row=2, max_row=num_rows + 1, max_col=num_cols), start=1):
            csv_file_path = os.path.join(self.csv_directory, f"{index}.csv")

            if os.path.exists(csv_file_path):
                is_empty = os.path.getsize(csv_file_path) == 0
                has_only_pdfs, pdf_urls = CSVHandler.has_only_pdf_urls(csv_file_path)

                if is_empty:
                    self._fill_row(row, RED_FILL)
                elif has_only_pdfs:
                    self._add_to_row(sheet, index, num_cols, pdf_urls)
                    self._fill_row(row, YELLOW_FILL)
                else:
                    self._fill_row(row, GREEN_FILL)
            else:
                self._fill_row(row, RED_FILL)

    def process_excel(self):
        '''
        Process an Excel file and create a result.xlsx file
        :return:
        '''
        df = pd.read_excel(self.input_file, sheet_name=0)
        num_rows, num_cols = df.shape

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'result'

        # Copy headers and data from the original DataFrame to the new sheet
        for r_idx, row in enumerate(df.itertuples(index=False), start=1):
            for c_idx, value in enumerate(row, start=1):
                if r_idx == 1:  # If it's the first row, copy headers
                    sheet.cell(row=1, column=c_idx, value=df.columns[c_idx - 1])
                sheet.cell(row=r_idx + 1, column=c_idx, value=value)

        self._process_csv_files(sheet, num_rows, num_cols)
        workbook.save(self.output_file)


if __name__ == '__main__':
    excel_processor = ExcelProcessor(
        input_file='resources/ifm_10.xlsx',
        output_file='result.xlsx',
        csv_directory='temp_files'
    )
    excel_processor.process_excel()
